from openpyxl import Workbook
from fpdf import FPDF 

class ExcelExporter:
    def __init__(self, filename):
        self.filename = filename

    def save_to_excel(self, data):
        workbook = Workbook()
        sheet = workbook.active

        # Write header
        sheet["A1"] = "Cake Name"
        sheet["B1"] = "Quantity Sold"

        row = 2
        for cake_name, quantity in data.items():
            sheet[f"A{row}"] = cake_name
            sheet[f"B{row}"] = quantity
            row += 1

        workbook.save(self.filename)
        print(f"Excel file saved as {self.filename}")


class PdfExporter:
    def __init__(self, filename):
        self.filename = filename

    def save_to_pdf(self, order):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        # Add order details
        pdf.cell(200, 10, txt=f"Invoice for {order.customer.name}", ln=True)
        pdf.ln(10)

        total_amount = 0
        for cake in order.cakes:
            pdf.cell(200, 10, txt=f"- {cake.name}: ${cake.price}", ln=True)
            total_amount += cake.price

        pdf.ln(10)
        pdf.cell(200, 10, txt=f"Total amount: ${total_amount}", ln=True)

        pdf.output(self.filename)
        print(f"PDF file saved as {self.filename}")


class PromoCode:
    def __init__(self, status: bool, promocode: str, percent: float):
        self.status = status
        self.promocode = promocode
        self.percent = percent

    def apply_promocode(self, price: float ):
        if self.status:
            discount = (self.percent / 100) * price
            new_price = price - discount
            return new_price
        else:
            return price
        
user_promocode = input("Please give promocode: ").capitalize().replace(" ", "")

promo = PromoCode(True, user_promocode, 20)
new_price = promo.apply_promocode(25)
print(new_price)

